"""
app.py — NMC Check Service (NextStep SaaS)
Embedded queue + dispatcher version.
"""

import os
import sys
import asyncio
import anyio
import logging
import json
import io
import shutil
import uuid
import datetime
import threading
import time
from pathlib import Path
from typing import Dict, Any, List, Optional
from functools import partial

if sys.platform.startswith("win"):
    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(name)s — %(message)s")
log = logging.getLogger(__name__)

from fastapi import FastAPI, UploadFile, File, HTTPException, Request
from fastapi.responses import HTMLResponse, JSONResponse, FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.background import BackgroundTask
from starlette.middleware.base import BaseHTTPMiddleware

from nmc_extract import extract_nmc_pin_from_bytes, parse_csv_pins, parse_xlsx_pins

APP_DIR = os.path.dirname(os.path.abspath(__file__))
REDIS_URL = os.environ.get("REDIS_URL", "redis://localhost:6379")
STORAGE_ROOT = Path("/tmp/nextstep")
STORAGE_ROOT.mkdir(parents=True, exist_ok=True)
BACKEND_VALIDATE_URL = os.environ.get("BACKEND_VALIDATE_URL", "https://nextstep-backend-e75l.onrender.com/api/validate-session")
APP_DASHBOARD_URL = os.environ.get("APP_DASHBOARD_URL", "https://nextstep-backend-e75l.onrender.com/dashboard")
APP_LOGIN_URL = os.environ.get("APP_LOGIN_URL", "https://nextstep-backend-e75l.onrender.com/login")
MAX_CONCURRENT_JOBS = max(1, int(os.environ.get("MAX_CONCURRENT_JOBS", "3")))
JOB_TTL_SECONDS = int(os.environ.get("NMC_JOB_TTL_SECONDS", str(60 * 60 * 24 * 2)))
DISPATCHER_ENABLED = os.environ.get("NMC_ENABLE_EMBEDDED_WORKER", "true").lower() in {"1", "true", "yes", "on"}
DISPATCHER_POLL_SECONDS = float(os.environ.get("NMC_DISPATCHER_POLL_SECONDS", "2"))

REDIS_STATE_PREFIX = "nextstep:nmc:job:"
REDIS_OWNER_PREFIX = "nextstep:nmc:owner:"
REDIS_PAYLOAD_PREFIX = "nextstep:nmc:payload:"
REDIS_PENDING_ZSET = "nextstep:nmc:pending"
REDIS_ACTIVE_SET = "nextstep:nmc:active"
REDIS_CLAIM_PREFIX = "nextstep:nmc:claim:"


class PersistNextStepTokenMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request, call_next):
        response = await call_next(request)
        token = request.query_params.get("ns_token")
        if token:
            response.set_cookie(
                key="ns_token",
                value=token,
                httponly=True,
                samesite="lax",
                secure=True,
                max_age=60 * 60 * 8,
            )
        return response


app = FastAPI(title="NMC Check — NextStep")
app.add_middleware(PersistNextStepTokenMiddleware)
app.mount("/static", StaticFiles(directory=os.path.join(APP_DIR, "static")), name="static")
templates = Jinja2Templates(directory=os.path.join(APP_DIR, "templates"))

_redis = None
_dispatcher_started = False
_dispatcher_lock = threading.Lock()


def get_redis():
    global _redis
    if _redis is None:
        try:
            import redis
            _redis = redis.Redis.from_url(REDIS_URL, decode_responses=False)
            _redis.ping()
            log.info("[Redis] Connected")
        except Exception as e:
            log.error("[Redis] %s", e)
            _redis = None
    return _redis


def _json_loads(raw):
    if not raw:
        return None
    if isinstance(raw, bytes):
        raw = raw.decode("utf-8")
    return json.loads(raw)


def _json_dumps(obj):
    return json.dumps(obj, separators=(",", ":"), ensure_ascii=False)


def _jget(job_id):
    r = get_redis()
    if not r:
        return None
    try:
        return _json_loads(r.get(f"{REDIS_STATE_PREFIX}{job_id}"))
    except Exception:
        return None


def _jset(job_id, state):
    r = get_redis()
    if not r:
        return
    try:
        r.setex(f"{REDIS_STATE_PREFIX}{job_id}", JOB_TTL_SECONDS, _json_dumps(state))
    except Exception as e:
        log.warning("[Redis] state set failed for %s: %s", job_id, e)


def _owner_set(job_id, tenant_id):
    r = get_redis()
    if not r:
        return
    try:
        r.setex(f"{REDIS_OWNER_PREFIX}{job_id}", JOB_TTL_SECONDS, str(tenant_id))
    except Exception:
        pass


def _owner_get(job_id):
    r = get_redis()
    if not r:
        return None
    try:
        v = r.get(f"{REDIS_OWNER_PREFIX}{job_id}")
        return int(v) if v else None
    except Exception:
        return None


def _payload_set(job_id: str, payload: Dict[str, Any]) -> None:
    r = get_redis()
    if not r:
        return
    try:
        r.setex(f"{REDIS_PAYLOAD_PREFIX}{job_id}", JOB_TTL_SECONDS, _json_dumps(payload))
    except Exception as e:
        log.error("[Queue] payload set failed for %s: %s", job_id, e)
        raise


def _payload_get(job_id: str) -> Optional[Dict[str, Any]]:
    r = get_redis()
    if not r:
        return None
    try:
        return _json_loads(r.get(f"{REDIS_PAYLOAD_PREFIX}{job_id}"))
    except Exception:
        return None


def _payload_delete(job_id: str) -> None:
    r = get_redis()
    if not r:
        return
    try:
        r.delete(f"{REDIS_PAYLOAD_PREFIX}{job_id}")
    except Exception:
        pass


def _queue_add(job_id: str) -> None:
    r = get_redis()
    if not r:
        raise RuntimeError("Redis is required for queueing jobs")
    score = time.time()
    r.zadd(REDIS_PENDING_ZSET, {job_id: score})


def _queue_remove(job_id: str) -> None:
    r = get_redis()
    if not r:
        return
    try:
        r.zrem(REDIS_PENDING_ZSET, job_id)
    except Exception:
        pass


def _active_add(job_id: str) -> None:
    r = get_redis()
    if not r:
        return
    r.sadd(REDIS_ACTIVE_SET, job_id)


def _active_remove(job_id: str) -> None:
    r = get_redis()
    if not r:
        return
    try:
        r.srem(REDIS_ACTIVE_SET, job_id)
        r.delete(f"{REDIS_CLAIM_PREFIX}{job_id}")
    except Exception:
        pass


def _active_count() -> int:
    r = get_redis()
    if not r:
        return 0
    try:
        return int(r.scard(REDIS_ACTIVE_SET) or 0)
    except Exception:
        return 0


def _acquire_job_claim(job_id: str) -> bool:
    r = get_redis()
    if not r:
        return False
    try:
        return bool(r.set(f"{REDIS_CLAIM_PREFIX}{job_id}", b"1", nx=True, ex=60 * 60 * 6))
    except Exception:
        return False


def _reset_embedded_worker_state() -> None:
    r = get_redis()
    if not r:
        return
    try:
        r.delete(REDIS_ACTIVE_SET)
        for raw in r.scan_iter(match=f"{REDIS_CLAIM_PREFIX}*"):
            r.delete(raw)
        for raw in r.scan_iter(match=f"{REDIS_PAYLOAD_PREFIX}*"):
            key = raw.decode("utf-8") if isinstance(raw, bytes) else raw
            job_id = key.split(REDIS_PAYLOAD_PREFIX, 1)[1]
            state = _jget(job_id) or {}
            if state.get("state") not in {"done", "completed", "failed"}:
                state["state"] = "queued"
                msg = state.get("message") or "Job queued — processing starts shortly..."
                if "recover" not in msg.lower():
                    state["message"] = "Job recovered after restart — processing starts shortly..."
                _jset(job_id, state)
                _queue_add(job_id)
    except Exception as e:
        log.warning("[Dispatcher] recovery reset failed: %s", e)


def _launch_background_job(job_id: str) -> None:
    t = threading.Thread(target=_run_job_wrapper, args=(job_id,), daemon=True, name=f"nmc-job-{job_id[:8]}")
    t.start()


def _run_job_wrapper(job_id: str) -> None:
    payload = _payload_get(job_id)
    if not payload:
        _active_remove(job_id)
        _queue_remove(job_id)
        state = _jget(job_id) or {}
        state.update({"state": "failed", "message": "Queued job payload expired before processing started.", "zip_ready": False})
        _jset(job_id, state)
        return

    try:
        import nmc_tasks
        log.info("[Dispatcher] Starting job %s", job_id)
        nmc_tasks.process_nmc_job(payload)
    except Exception as e:
        log.exception("[Dispatcher] Job %s crashed", job_id)
        state = _jget(job_id) or {}
        rows = state.get("rows") or []
        state.update({
            "state": "failed",
            "rows": rows,
            "zip_ready": False,
            "message": f"Job failed: {e}",
        })
        _jset(job_id, state)
        try:
            import db
            db_job_id = payload.get("db_job_id")
            if db_job_id:
                done = sum(1 for row in rows if row.get("status") == "done")
                failed = sum(1 for row in rows if row.get("status") == "failed")
                db.update_job_status(db_job_id=db_job_id, status="failed", successful_items=done, failed_items=failed)
        except Exception:
            log.exception("[Dispatcher] Failed to mark DB job failed for %s", job_id)
    finally:
        _active_remove(job_id)
        _queue_remove(job_id)
        _payload_delete(job_id)


def _dispatcher_loop() -> None:
    while True:
        try:
            r = get_redis()
            if not r:
                time.sleep(max(5.0, DISPATCHER_POLL_SECONDS))
                continue

            active = _active_count()
            if active >= MAX_CONCURRENT_JOBS:
                time.sleep(DISPATCHER_POLL_SECONDS)
                continue

            slots = max(0, MAX_CONCURRENT_JOBS - active)
            candidates = r.zrange(REDIS_PENDING_ZSET, 0, max(0, slots * 4 - 1))
            if not candidates:
                time.sleep(DISPATCHER_POLL_SECONDS)
                continue

            launched = 0
            for raw_job_id in candidates:
                if active + launched >= MAX_CONCURRENT_JOBS:
                    break
                job_id = raw_job_id.decode("utf-8") if isinstance(raw_job_id, bytes) else raw_job_id
                state = _jget(job_id) or {}
                if state.get("state") in {"done", "completed", "failed"}:
                    _queue_remove(job_id)
                    _payload_delete(job_id)
                    continue
                if not _payload_get(job_id):
                    _queue_remove(job_id)
                    state.update({"state": "failed", "message": "Queued job payload was missing.", "zip_ready": False})
                    _jset(job_id, state)
                    continue
                if not _acquire_job_claim(job_id):
                    continue
                _active_add(job_id)
                _queue_remove(job_id)
                _launch_background_job(job_id)
                launched += 1

            time.sleep(DISPATCHER_POLL_SECONDS)
        except Exception as e:
            log.exception("[Dispatcher] loop error: %s", e)
            time.sleep(max(5.0, DISPATCHER_POLL_SECONDS))


def _start_dispatcher_once() -> None:
    global _dispatcher_started
    if not DISPATCHER_ENABLED:
        log.info("[Dispatcher] Embedded dispatcher disabled")
        return
    with _dispatcher_lock:
        if _dispatcher_started:
            return
        get_redis()
        _reset_embedded_worker_state()
        t = threading.Thread(target=_dispatcher_loop, daemon=True, name="nmc-dispatcher")
        t.start()
        _dispatcher_started = True
        log.info("[Dispatcher] Embedded dispatcher started (max_concurrent=%s)", MAX_CONCURRENT_JOBS)


@app.on_event("startup")
def _startup() -> None:
    _start_dispatcher_once()


def _validate_via_backend(token: str):
    if not token:
        return None
    try:
        import requests
        resp = requests.get(BACKEND_VALIDATE_URL, params={"token": token}, timeout=8)
        if resp.status_code != 200:
            return None
        data = resp.json()
        if not data.get("valid"):
            return None
        user = data.get("user") or {}
        tenant = data.get("tenant") or {}
        return {
            "user_id": user.get("id"),
            "tenant_id": tenant.get("id"),
            "role": user.get("role", "admin"),
            "email": user.get("email"),
            "name": user.get("name"),
        }
    except Exception as e:
        log.warning("[Auth backend] %s", e)
        return None


def _get_ctx(request: Request):
    token = (
        request.headers.get("X-NextStep-Token")
        or request.cookies.get("ns_token")
        or request.query_params.get("ns_token")
        or ""
    )
    if not token:
        return None
    ctx = _validate_via_backend(token)
    if ctx:
        return ctx
    try:
        import db
        return db.validate_user_token(token)
    except Exception as e:
        log.warning("[Auth db] %s", e)
        return None


def _auth(request: Request):
    ctx = _get_ctx(request)
    if not ctx:
        raise HTTPException(status_code=401, detail=f"Not authenticated. Please log in at {APP_LOGIN_URL}")
    return ctx


def _storage(tenant_id, user_id, job_id):
    p = STORAGE_ROOT / str(tenant_id) / str(user_id) / job_id
    p.mkdir(parents=True, exist_ok=True)
    return p


def _create_rows(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    rows = []
    for i, it in enumerate(items[:100]):
        rows.append({
            "row": i + 1,
            "status": "queued",
            "nmc_pin": (it.get("nmc_pin") if isinstance(it, dict) else "") or "",
            "original_filename": (it.get("original_filename") if isinstance(it, dict) else f"Row {i+1}"),
            "name": "",
            "expiry_date": "",
            "status_text": "",
            "pdf_filename": "",
            "pdf_url": "",
            "error": "",
        })
    return rows


def _enqueue_job(job_id: str, state: Dict[str, Any], payload: Dict[str, Any], tenant_id: int) -> None:
    _jset(job_id, state)
    _owner_set(job_id, tenant_id)
    _payload_set(job_id, payload)
    _queue_add(job_id)


@app.get("/", response_class=HTMLResponse)
def home(request: Request):
    response = templates.TemplateResponse(
        "index.html",
        {"request": request, "dashboard_url": APP_DASHBOARD_URL, "login_url": APP_LOGIN_URL},
    )
    token = request.query_params.get("ns_token")
    if token:
        response.set_cookie(
            key="ns_token",
            value=token,
            httponly=True,
            samesite="lax",
            secure=True,
            max_age=60 * 60 * 8,
        )
    return response


@app.get("/health")
def health():
    r = get_redis()
    ok = False
    try:
        if r:
            r.ping()
            ok = True
    except Exception:
        pass
    return {
        "ok": True,
        "redis": ok,
        "twocaptcha": bool(os.getenv("TWOCAPTCHA_API_KEY")),
        "db": bool(os.getenv("DATABASE_URL")),
        "embedded_worker": DISPATCHER_ENABLED,
        "max_concurrent_jobs": MAX_CONCURRENT_JOBS,
        "active_jobs": _active_count() if ok else 0,
    }


@app.post("/nmc/extract")
async def nmc_extract(request: Request, files: List[UploadFile] = File(...)):
    _auth(request)
    if not files:
        raise HTTPException(400, "No file uploaded.")
    items: List[Dict] = []
    for file in files[:100]:
        content = await file.read()
        fname = file.filename or "upload"
        flower = fname.lower()
        if len(content) > 25 * 1024 * 1024:
            raise HTTPException(413, f"'{fname}' too large.")
        if flower.endswith(".csv"):
            for r in parse_csv_pins(content):
                if len(items) >= 100:
                    break
                items.append({
                    "original_filename": f"{fname} ({r['original_filename']})",
                    "nmc_pin": r["nmc_pin"],
                    "confidence": 95,
                    "source": "Spreadsheet",
                })
            continue
        if flower.endswith(".xlsx"):
            for r in parse_xlsx_pins(content):
                if len(items) >= 100:
                    break
                items.append({
                    "original_filename": f"{fname} ({r['original_filename']})",
                    "nmc_pin": r["nmc_pin"],
                    "confidence": 95,
                    "source": "Spreadsheet",
                })
            continue
        if flower.endswith(".webp"):
            try:
                from PIL import Image
                im = Image.open(io.BytesIO(content)).convert("RGB")
                buf = io.BytesIO()
                im.save(buf, "PNG")
                content = buf.getvalue()
                fname = fname[:-5] + ".png"
                flower = fname.lower()
            except Exception:
                pass
        result = extract_nmc_pin_from_bytes(content, fname)
        pin = (result.get("nmc_pin") or "").strip().upper()
        conf = int((result.get("confidence") or {}).get("nmc_pin", 0) * 100)
        items.append({
            "original_filename": fname,
            "nmc_pin": pin,
            "confidence": conf,
            "source": "PDF text" if flower.endswith(".pdf") else "Image scan",
        })
        if len(items) >= 100:
            break
    return JSONResponse({"items": items})


@app.post("/nmc/run")
async def nmc_run(request: Request):
    ctx = _auth(request)
    tenant_id = ctx["tenant_id"]
    user_id = ctx["user_id"]

    payload: Dict = {}
    ctype = (request.headers.get("content-type") or "").lower()
    if "application/json" in ctype:
        try:
            payload = await request.json()
        except Exception:
            payload = {}
    else:
        form = await request.form()
        payload = dict(form)

    items = payload.get("items")
    if not isinstance(items, list) or not items:
        raise HTTPException(400, "No items provided.")
    items = items[:100]

    try:
        import db
        tokens = db.get_tenant_tokens_remaining(tenant_id)
        if tokens == 0:
            raise HTTPException(402, "No tokens remaining. Please contact NextStep to top up.")
        if 0 < tokens < len(items):
            items = items[:tokens]
    except HTTPException:
        raise
    except Exception as e:
        log.warning("[Run] Token check skipped: %s", e)

    if not items:
        raise HTTPException(400, "No items remaining after token check.")

    job_id = str(uuid.uuid4())
    storage_path = _storage(tenant_id, user_id, job_id)

    db_job_id = None
    try:
        import db
        db_job_id = db.create_job_record(tenant_id=tenant_id, user_id=user_id, total_items=len(items))
    except Exception as e:
        log.warning("[Run] DB record failed: %s", e)

    rows = _create_rows(items)
    state = {
        "state": "queued",
        "rows": rows,
        "zip_ready": False,
        "zip_name": "",
        "zip_url": "",
        "message": "Job queued — processing starts shortly...",
        "successful": 0,
        "failed": 0,
    }
    job_payload = {
        "job_id": job_id,
        "db_job_id": db_job_id,
        "tenant_id": tenant_id,
        "user_id": user_id,
        "items": items,
        "storage_path": str(storage_path),
    }

    _enqueue_job(job_id, state, job_payload, tenant_id)
    log.info("[Run] Job %s queued %d items tenant=%d", job_id, len(items), tenant_id)
    return JSONResponse({"job_id": job_id, "status_url": f"/nmc/status/{job_id}", "rows": rows, "queued": True})


@app.get("/nmc/status/{job_id}")
async def nmc_status(job_id: str, request: Request):
    _auth(request)
    state = _jget(job_id)
    if not state:
        raise HTTPException(404, "Job expired or not found.")
    rows = state.get("rows") or []
    done = sum(1 for r in rows if r.get("status") in ("done", "failed"))
    return JSONResponse({
        "job_id": job_id,
        "state": state.get("state", "queued"),
        "running": {"done": done, "total": len(rows) or 1},
        "rows": rows,
        "zip_ready": bool(state.get("zip_ready")),
        "zip_name": state.get("zip_name", ""),
        "zip_url": state.get("zip_url", ""),
        "message": state.get("message", ""),
        "successful": state.get("successful", 0),
        "failed": state.get("failed", 0),
    })


@app.get("/nmc/download/{job_id}/{name}")
async def nmc_download(job_id: str, name: str, request: Request):
    ctx = _auth(request)
    tenant_id = ctx["tenant_id"]
    job_tenant = _owner_get(job_id)
    if job_tenant is not None and job_tenant != tenant_id:
        log.warning("[DL] Tenant %d tried job owned by %d", tenant_id, job_tenant)
        raise HTTPException(403, "Access denied.")
    tenant_root = STORAGE_ROOT / str(tenant_id)
    file_path = None
    for p in tenant_root.rglob(name):
        if job_id in str(p):
            file_path = p
            break
    if not file_path or not file_path.exists():
        raise HTTPException(404, "Download expired or not found.")
    bg = None
    if name.lower().endswith(".zip"):
        sp = STORAGE_ROOT / str(tenant_id) / str(ctx["user_id"]) / job_id
        bg = BackgroundTask(_cleanup, sp, job_id)
    return FileResponse(str(file_path), filename=name, media_type="application/octet-stream", background=bg)


def _cleanup(storage_path: Path, job_id: str):
    import time as t
    t.sleep(5)
    try:
        if storage_path.exists():
            shutil.rmtree(storage_path, ignore_errors=True)
    except Exception as e:
        log.warning("[Cleanup] %s", e)
    r = get_redis()
    if r:
        try:
            r.delete(f"{REDIS_STATE_PREFIX}{job_id}")
            r.delete(f"{REDIS_OWNER_PREFIX}{job_id}")
            r.delete(f"{REDIS_PAYLOAD_PREFIX}{job_id}")
            r.zrem(REDIS_PENDING_ZSET, job_id)
            r.srem(REDIS_ACTIVE_SET, job_id)
            r.delete(f"{REDIS_CLAIM_PREFIX}{job_id}")
        except Exception:
            pass


@app.get("/nmc/export/excel/{job_id}")
async def nmc_export_excel(job_id: str, request: Request):
    ctx = _auth(request)
    jt = _owner_get(job_id)
    if jt is not None and jt != ctx["tenant_id"]:
        raise HTTPException(403, "Access denied.")
    state = _jget(job_id)
    if not state:
        raise HTTPException(404, "Job not found.")
    rows = state.get("rows") or []

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "NMC Results"
    headers = ["#", "NMC PIN", "Name", "Expiry Date", "Status", "Result"]
    hf = PatternFill("solid", fgColor="1e3a8a")
    hfont = Font(bold=True, color="FFFFFF")
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = hf
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center")
    for w, col in zip([6, 16, 30, 14, 36, 12], "ABCDEF"):
        ws.column_dimensions[col].width = w
    for r in rows:
        rn = ws.max_row + 1
        st = r.get("status", "")
        rl = "Done" if st == "done" else ("Failed" if st == "failed" else st.title())
        ws.append([r.get("row", ""), r.get("nmc_pin", ""), r.get("name", ""), r.get("expiry_date", ""), r.get("status_text", ""), rl])
        rc = ws.cell(row=rn, column=6)
        if st == "done":
            rc.fill = PatternFill("solid", fgColor="d1fae5")
            rc.font = Font(color="065f46")
        elif st == "failed":
            rc.fill = PatternFill("solid", fgColor="fee2e2")
            rc.font = Font(color="991b1b")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    try:
        from zoneinfo import ZoneInfo
        ds = datetime.datetime.now(tz=ZoneInfo("Europe/London")).strftime("%d.%m.%Y")
    except Exception:
        ds = datetime.datetime.utcnow().strftime("%d.%m.%Y")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="NMC_Results_{ds}.xlsx"'},
    )


@app.get("/nmc/export/csv/{job_id}")
async def nmc_export_csv(job_id: str, request: Request):
    ctx = _auth(request)
    jt = _owner_get(job_id)
    if jt is not None and jt != ctx["tenant_id"]:
        raise HTTPException(403, "Access denied.")
    state = _jget(job_id)
    if not state:
        raise HTTPException(404, "Job not found.")
    rows = state.get("rows") or []
    import csv as cm
    buf = io.StringIO()
    w = cm.writer(buf)
    w.writerow(["#", "NMC PIN", "Name", "Expiry Date", "Status", "Result"])
    for r in rows:
        st = r.get("status", "")
        w.writerow([
            r.get("row", ""),
            r.get("nmc_pin", ""),
            r.get("name", ""),
            r.get("expiry_date", ""),
            r.get("status_text", ""),
            "Done" if st == "done" else ("Failed" if st == "failed" else st.title()),
        ])
    try:
        from zoneinfo import ZoneInfo
        ds = datetime.datetime.now(tz=ZoneInfo("Europe/London")).strftime("%d.%m.%Y")
    except Exception:
        ds = datetime.datetime.utcnow().strftime("%d.%m.%Y")
    return StreamingResponse(
        io.BytesIO(buf.getvalue().encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="NMC_Results_{ds}.csv"'},
    )


@app.post("/nmc/rerun")
async def nmc_rerun(request: Request):
    ctx = _auth(request)
    tenant_id = ctx["tenant_id"]
    user_id = ctx["user_id"]

    try:
        payload: Dict = await request.json()
    except Exception:
        raise HTTPException(400, "Invalid JSON.")

    old_job_id = (payload.get("job_id") or "").strip()
    dirty_items = payload.get("items") or []
    if not old_job_id:
        raise HTTPException(400, "job_id is required.")
    if not dirty_items:
        raise HTTPException(400, "No items provided for rerun.")

    job_tenant = _owner_get(old_job_id)
    if job_tenant is not None and job_tenant != tenant_id:
        raise HTTPException(403, "Access denied.")

    old_state = _jget(old_job_id)
    if not old_state:
        raise HTTPException(404, "Original job expired. Please run a fresh check.")

    old_rows = old_state.get("rows") or []
    old_storage = STORAGE_ROOT / str(tenant_id) / str(user_id) / old_job_id

    dirty_count = len(dirty_items)
    try:
        import db
        tokens = db.get_tenant_tokens_remaining(tenant_id)
        if tokens == 0:
            raise HTTPException(402, "No tokens remaining.")
        if 0 < tokens < dirty_count:
            dirty_items = dirty_items[:tokens]
            dirty_count = len(dirty_items)
    except HTTPException:
        raise
    except Exception as e:
        log.warning("[Rerun] Token check skipped: %s", e)

    if not dirty_items:
        raise HTTPException(400, "No items remaining after token check.")

    new_job_id = str(uuid.uuid4())
    storage_path = old_storage

    db_job_id = None
    try:
        import db
        db_job_id = db.create_job_record(tenant_id=tenant_id, user_id=user_id, total_items=dirty_count)
    except Exception as e:
        log.warning("[Rerun] DB record failed: %s", e)

    dirty_row_nums = {it.get("row_number") for it in dirty_items if it.get("row_number")}
    merged_rows = []
    for r in old_rows:
        rnum = r.get("row", 0)
        if rnum in dirty_row_nums:
            new_pin = next((it.get("nmc_pin", "") for it in dirty_items if it.get("row_number") == rnum), r.get("nmc_pin", ""))
            merged_rows.append({
                **r,
                "status": "queued",
                "nmc_pin": new_pin.strip().upper(),
                "pdf_filename": "",
                "pdf_url": "",
                "error": "",
                "name": "",
                "expiry_date": "",
                "status_text": "",
            })
        else:
            merged_rows.append(r)

    rerun_items = []
    for it in dirty_items:
        rerun_items.append({
            "nmc_pin": (it.get("nmc_pin") or "").strip().upper(),
            "original_filename": it.get("original_filename") or f"Row {it.get('row_number', '')}",
            "row_number": it.get("row_number"),
        })

    state = {
        "state": "queued",
        "rows": merged_rows,
        "zip_ready": old_state.get("zip_ready", False),
        "zip_name": old_state.get("zip_name", ""),
        "zip_url": old_state.get("zip_url", ""),
        "message": f"Rerunning {dirty_count} edited row(s)…",
        "successful": 0,
        "failed": 0,
    }
    job_payload = {
        "job_id": new_job_id,
        "db_job_id": db_job_id,
        "tenant_id": tenant_id,
        "user_id": user_id,
        "items": rerun_items,
        "storage_path": str(storage_path),
        "is_rerun": True,
        "old_rows": old_rows,
        "dirty_row_nums": list(dirty_row_nums),
    }

    _enqueue_job(new_job_id, state, job_payload, tenant_id)
    log.info("[Rerun] Job %s queued — %d dirty rows", new_job_id, dirty_count)
    return JSONResponse({"job_id": new_job_id, "status_url": f"/nmc/status/{new_job_id}", "rows": merged_rows, "queued": True})
