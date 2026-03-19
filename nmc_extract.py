import io
import os
import re
from pathlib import Path
from typing import Optional, Dict, Any, List, Tuple

import pdfplumber
import fitz  # PyMuPDF

try:
    from google import genai
    from google.genai import types
except Exception:
    genai = None
    types = None

# ------------------------------------------------------------
# NMC PIN format: YY M #### C
#   YY: 2 digits (year)
#   M : A–L (month code)
#   ####: 4 digits
#   C : country code in {E,S,W,N,O}
# Examples: 23B0365O, 09B0112E, 16J0151E, 18G3283E
# ------------------------------------------------------------

STRICT_NMC_RE = re.compile(r"\b\d{2}[A-L]\d{4}[ESWNO]\b", re.I)
LOOSE_8_RE = re.compile(r"\b\d{2}[A-Z]\d{4}[A-Z]\b", re.I)

ANCHOR_RE = re.compile(
    r"("
    r"NMC\s*PIN"
    r"|NMC\s*PIN\s*NUMBER"
    r"|PIN\s*NUMBER"
    r"|PIN\s*NO\.?"
    r"|PIN\s*#"
    r"|PIN\s*:"
    r"|REGISTRATION\s*NUMBER"
    r"|NMC\s*REGISTRATION\s*NUMBER"
    r"|PERSONAL\s*IDENTIFICATION\s*NUMBER"
    r")",
    re.I
)

GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "").strip()
GEMINI_MODEL_FAST = os.getenv("GEMINI_MODEL_FAST", "gemini-2.0-flash")
GEMINI_MODEL_STRONG = os.getenv("GEMINI_MODEL_STRONG", "gemini-2.5-pro")

NMC_PDF_TEXT_PAGES = int(os.getenv("NMC_PDF_TEXT_PAGES", "8"))
NMC_PDF_IMAGE_PAGES = int(os.getenv("NMC_PDF_IMAGE_PAGES", "4"))

_client = None
if GEMINI_API_KEY and genai is not None:
    try:
        _client = genai.Client(api_key=GEMINI_API_KEY)
    except Exception:
        _client = None


_DIGIT_FIX = {"O": "0", "I": "1", "L": "1", "S": "5", "B": "8"}
_LETTER_FIX = {"0": "O", "1": "I", "5": "S", "8": "B"}


def _normalize_token(token: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", (token or "").upper())


def _fix_by_position(token8: str) -> str:
    s = list(token8)
    for i in (0, 1, 3, 4, 5, 6):
        s[i] = _DIGIT_FIX.get(s[i], s[i])
    for i in (2, 7):
        s[i] = _LETTER_FIX.get(s[i], s[i])
    return "".join(s)


def _validate_strict(pin: str) -> bool:
    return bool(STRICT_NMC_RE.fullmatch(pin))


def _clean_and_validate(raw: str) -> Optional[str]:
    s = _normalize_token(raw)
    if not s:
        return None
    m = STRICT_NMC_RE.search(s)
    if m:
        return m.group(0).upper()
    if len(s) >= 8:
        for start in range(0, min(len(s) - 7, 16)):
            chunk = s[start:start + 8]
            if len(chunk) != 8:
                continue
            fixed = _fix_by_position(chunk)
            if _validate_strict(fixed):
                return fixed
    m2 = LOOSE_8_RE.search(s)
    if m2:
        fixed = _fix_by_position(m2.group(0).upper())
        if _validate_strict(fixed):
            return fixed
    return None


def _read_pdf_text(path: Path, max_pages: int = 3) -> str:
    try:
        with pdfplumber.open(str(path)) as pdf:
            parts = []
            for page in pdf.pages[:max_pages]:
                t = page.extract_text() or ""
                if t:
                    parts.append(t)
            return "\n".join(parts)
    except Exception:
        return ""


def _pdf_to_images(path: Path, max_pages: int = 2) -> List[Tuple[bytes, str]]:
    out: List[Tuple[bytes, str]] = []
    try:
        doc = fitz.open(str(path))
        for i in range(min(max_pages, doc.page_count)):
            page = doc.load_page(i)
            pix = page.get_pixmap(dpi=200)
            out.append((pix.tobytes("png"), "image/png"))
        doc.close()
    except Exception:
        pass
    return out


def _file_to_image(path: Path) -> Optional[Tuple[bytes, str]]:
    ext = path.suffix.lower()
    mime = {
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".webp": "image/webp",
    }.get(ext)
    if not mime:
        return None
    try:
        return (path.read_bytes(), mime)
    except Exception:
        return None


def _extract_from_text(text: str) -> Tuple[Optional[str], float]:
    if not text:
        return None, 0.0

    T = text.upper()

    for m in ANCHOR_RE.finditer(T):
        window = T[m.end(): m.end() + 120]
        m_strict = STRICT_NMC_RE.search(window)
        if m_strict:
            return m_strict.group(0).upper(), 0.99
        m_loose = LOOSE_8_RE.search(window)
        if m_loose:
            pin = _clean_and_validate(m_loose.group(0))
            if pin:
                return pin, 0.98
        tokenish = re.findall(r"[A-Z0-9]{7,12}", window)
        for tok in tokenish[:3]:
            pin = _clean_and_validate(tok)
            if pin:
                return pin, 0.96

    m2 = STRICT_NMC_RE.search(T)
    if m2:
        return m2.group(0).upper(), 0.95

    candidates = LOOSE_8_RE.findall(T)
    if candidates:
        for cand in candidates:
            idx = T.find(cand.upper())
            if idx != -1:
                vicinity = T[max(0, idx - 80): idx + 80]
                if "NMC" in vicinity:
                    pin = _clean_and_validate(cand)
                    if pin:
                        return pin, 0.92
        for cand in candidates:
            pin = _clean_and_validate(cand)
            if pin:
                return pin, 0.88

    return None, 0.0


def _gemini_extract(images: List[Tuple[bytes, str]]) -> Tuple[Optional[str], float]:
    if _client is None or types is None or not images:
        return None, 0.0

    prompt = (
        "Extract the NMC PIN from the document image. Look for labels like 'NMC PIN', 'PIN number', or 'Registration number'. "
        "Return ONLY the PIN value, nothing else.\n"
        "Valid format:\n"
        "- 2 digits (year)\n"
        "- 1 letter A to L (month code)\n"
        "- 4 digits\n"
        "- 1 letter (country code: E, S, W, N, or O)\n"
        "Example: 12A3456S"
    )

    def _call(model: str) -> str:
        parts = [types.Part.from_text(text=prompt)]
        for b, mime in images:
            parts.append(types.Part.from_bytes(data=b, mime_type=mime))
        resp = _client.models.generate_content(
            model=model,
            contents=[types.Content(role="user", parts=parts)],
        )
        return (getattr(resp, "text", None) or "").strip()

    for model in (GEMINI_MODEL_FAST, GEMINI_MODEL_STRONG):
        try:
            txt = _call(model)
            pin = _clean_and_validate(txt)
            if pin:
                return pin, 0.90 if model == GEMINI_MODEL_FAST else 0.93
        except Exception:
            continue
    return None, 0.0


def _images_from_bytes(content: bytes, filename: str) -> List[Tuple[bytes, str]]:
    """Convert various file types to image bytes for Gemini."""
    fname = filename.lower()

    if fname.endswith(".webp"):
        try:
            from PIL import Image
            im = Image.open(io.BytesIO(content)).convert("RGB")
            buf = io.BytesIO()
            im.save(buf, format="PNG")
            return [(buf.getvalue(), "image/png")]
        except Exception:
            return []

    ext_mime = {
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
    }
    for ext, mime in ext_mime.items():
        if fname.endswith(ext):
            return [(content, mime)]

    return []


def extract_nmc_pin(file_path: Path) -> Dict[str, Any]:
    path = Path(file_path)

    if path.suffix.lower() == ".pdf":
        try:
            with pdfplumber.open(str(path)) as pdf:
                combined_parts = []
                for page in pdf.pages[:max(1, NMC_PDF_TEXT_PAGES)]:
                    t = page.extract_text() or ""
                    if t:
                        pin_page, conf_page = _extract_from_text(t)
                        if pin_page:
                            return {"ok": True, "nmc_pin": pin_page, "confidence": {"nmc_pin": conf_page}}
                        combined_parts.append(t)
                text = "\n".join(combined_parts)
        except Exception:
            text = ""

        pin, conf = _extract_from_text(text)
        if pin:
            return {"ok": True, "nmc_pin": pin, "confidence": {"nmc_pin": conf}}

        imgs = _pdf_to_images(path, max_pages=max(1, NMC_PDF_IMAGE_PAGES))
        pin2, conf2 = _gemini_extract(imgs)
        if pin2:
            return {"ok": True, "nmc_pin": pin2, "confidence": {"nmc_pin": conf2}}

        return {"ok": False, "nmc_pin": None, "confidence": {"nmc_pin": 0.0}}

    img = _file_to_image(path)
    if img:
        pin3, conf3 = _gemini_extract([img])
        if pin3:
            return {"ok": True, "nmc_pin": pin3, "confidence": {"nmc_pin": conf3}}
        return {"ok": False, "nmc_pin": None, "confidence": {"nmc_pin": 0.0}}

    try:
        text = path.read_text(errors="ignore")
    except Exception:
        text = ""
    pin4, conf4 = _extract_from_text(text)
    return {"ok": bool(pin4), "nmc_pin": pin4, "confidence": {"nmc_pin": conf4}}


def extract_nmc_pin_from_bytes(content: bytes, filename: str) -> Dict[str, Any]:
    """Extract NMC PIN directly from bytes without writing to disk."""
    fname = filename.lower()

    if fname.endswith(".pdf"):
        try:
            with pdfplumber.open(io.BytesIO(content)) as pdf:
                combined_parts = []
                for page in pdf.pages[:max(1, NMC_PDF_TEXT_PAGES)]:
                    t = page.extract_text() or ""
                    if t:
                        pin_page, conf_page = _extract_from_text(t)
                        if pin_page:
                            return {"ok": True, "nmc_pin": pin_page, "confidence": {"nmc_pin": conf_page}}
                        combined_parts.append(t)
                text = "\n".join(combined_parts)
        except Exception:
            text = ""

        pin, conf = _extract_from_text(text)
        if pin:
            return {"ok": True, "nmc_pin": pin, "confidence": {"nmc_pin": conf}}

        try:
            doc = fitz.open(stream=content, filetype="pdf")
            imgs = []
            for i in range(min(NMC_PDF_IMAGE_PAGES, doc.page_count)):
                page = doc.load_page(i)
                pix = page.get_pixmap(dpi=200)
                imgs.append((pix.tobytes("png"), "image/png"))
            doc.close()
        except Exception:
            imgs = []

        pin2, conf2 = _gemini_extract(imgs)
        if pin2:
            return {"ok": True, "nmc_pin": pin2, "confidence": {"nmc_pin": conf2}}

        return {"ok": False, "nmc_pin": None, "confidence": {"nmc_pin": 0.0}}

    imgs = _images_from_bytes(content, filename)
    if imgs:
        pin3, conf3 = _gemini_extract(imgs)
        if pin3:
            return {"ok": True, "nmc_pin": pin3, "confidence": {"nmc_pin": conf3}}
        return {"ok": False, "nmc_pin": None, "confidence": {"nmc_pin": 0.0}}

    try:
        text = content.decode("utf-8", errors="ignore")
    except Exception:
        text = ""
    pin4, conf4 = _extract_from_text(text)
    return {"ok": bool(pin4), "nmc_pin": pin4, "confidence": {"nmc_pin": conf4}}


def parse_csv_pins(content: bytes) -> List[Dict[str, Any]]:
    """Parse CSV/XLSX with NMC PIN column. Returns list of {nmc_pin, row}."""
    import csv
    rows = []
    text = None
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = content.decode(enc)
            break
        except Exception:
            continue
    if not text:
        return []

    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        return []

    pin_col = None
    for f in reader.fieldnames:
        norm = re.sub(r"[^a-z0-9]", "", f.lower())
        if norm in ("nmcpin", "pin", "nmcpinnumber", "pinnumber", "registrationnumber"):
            pin_col = f
            break

    for i, row in enumerate(reader, start=2):
        if pin_col:
            raw = (row.get(pin_col) or "").strip()
        else:
            raw = next((v for v in row.values() if v and _clean_and_validate(str(v).strip())), "")

        pin = _clean_and_validate(raw)
        if pin:
            rows.append({"nmc_pin": pin, "row": i, "original_filename": f"Row {i}"})
    return rows


def parse_xlsx_pins(content: bytes) -> List[Dict[str, Any]]:
    """Parse XLSX with NMC PIN column."""
    try:
        from openpyxl import load_workbook
    except Exception:
        return []

    rows = []
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return []

    headers = [str(h).strip() if h is not None else "" for h in all_rows[0]]
    pin_col_idx = None
    for i, h in enumerate(headers):
        norm = re.sub(r"[^a-z0-9]", "", h.lower())
        if norm in ("nmcpin", "pin", "nmcpinnumber", "pinnumber", "registrationnumber"):
            pin_col_idx = i
            break

    for row_idx, row in enumerate(all_rows[1:], start=2):
        if pin_col_idx is not None and pin_col_idx < len(row):
            raw = str(row[pin_col_idx] or "").strip()
        else:
            raw = next((str(v) for v in row if v and _clean_and_validate(str(v).strip())), "")
        pin = _clean_and_validate(raw)
        if pin:
            rows.append({"nmc_pin": pin, "row": row_idx, "original_filename": f"Row {row_idx}"})
    return rows
